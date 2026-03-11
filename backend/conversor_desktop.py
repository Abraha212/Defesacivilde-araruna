"""
╔══════════════════════════════════════════════════════════════════╗
║   CONVERSOR NETCDF  v3.0                                         ║
║                                                                  ║
║   By: Offnen Soluções e Desenvolvimento                          ║
║                                                                  ║
║   Formatos de saída: CSV  |  Excel (XLSX)  |  XML                ║
║   Suporta arquivos de 3 GB, 4 GB ou mais                         ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import gc
import threading
import subprocess
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from pathlib import Path

# ── AUTO-INSTALAÇÃO DE DEPENDÊNCIAS ─────────────────────────────────────────
_DEPS = [
    ('xarray',   'xarray>=2024.1.0'),
    ('netCDF4',  'netCDF4>=1.6.0'),
    ('pandas',   'pandas>=2.0.0'),
    ('numpy',    'numpy>=1.24.0'),
    ('openpyxl', 'openpyxl>=3.1.0'),
]

def _instalar_deps():
    """Verifica e instala dependências ausentes automaticamente."""
    ausentes = []
    for modulo, pacote in _DEPS:
        try:
            __import__(modulo)
        except ImportError:
            ausentes.append(pacote)

    if not ausentes:
        return True  # Tudo ok

    # Janela de instalação
    splash = tk.Tk()
    splash.title('Configurando...')
    splash.configure(bg='#0e1520')
    splash.resizable(False, False)
    splash.geometry('480x200')
    try:
        splash.eval('tk::PlaceWindow . center')
    except Exception:
        pass

    tk.Label(splash, text='Instalando dependencias...',
             font=('Segoe UI', 13, 'bold'),
             bg='#0e1520', fg='#e87722').pack(pady=(30, 6))

    lbl = tk.Label(splash, text='Aguarde, isso acontece so na primeira vez.',
                   font=('Segoe UI', 9), bg='#0e1520', fg='#64748b')
    lbl.pack()

    bar = ttk.Progressbar(splash, mode='indeterminate', length=380)
    bar.pack(pady=20)
    bar.start(12)
    splash.update()

    try:
        subprocess.check_call(
            [sys.executable, '-m', 'pip', 'install', '--quiet'] + ausentes,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        bar.stop()
        splash.destroy()
        return True
    except subprocess.CalledProcessError:
        bar.stop()
        splash.destroy()
        messagebox.showerror(
            'Erro de instalacao',
            'Nao foi possivel instalar as dependencias automaticamente.\n\n'
            'Execute manualmente no terminal:\n'
            f'pip install {" ".join(ausentes)}'
        )
        return False

# Só executa a verificação quando rodando como script Python (não no .exe)
if not getattr(sys, 'frozen', False):
    _root_tmp = tk.Tk()
    _root_tmp.withdraw()
    if not _instalar_deps():
        sys.exit(1)
    _root_tmp.destroy()

# ── DEPENDÊNCIAS EXTERNAS ────────────────────────────────────────────────────
try:
    import xarray as xr
    import pandas as pd
    import numpy as np
    LIBS_OK = True
except ImportError as e:
    LIBS_OK = False
    MISSING_LIB = str(e)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl import load_workbook
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── PALETA DE CORES ──────────────────────────────────────────────────────────
C = {
    'bg':        '#080d14',   # Fundo principal (muito escuro)
    'bg2':       '#0e1520',   # Painel lateral
    'card':      '#141e2e',   # Cards
    'card2':     '#1a2640',   # Cards secundários
    'accent':    '#e87722',   # Laranja (cor da prefeitura)
    'accent_h':  '#f59e0b',   # Laranja hover
    'blue':      '#3b82f6',   # Azul
    'green':     '#22c55e',   # Verde
    'red':       '#ef4444',   # Vermelho
    'yellow':    '#f59e0b',   # Amarelo
    'purple':    '#8b5cf6',   # Roxo
    'txt':       '#e2e8f0',   # Texto principal
    'txt2':      '#64748b',   # Texto secundário
    'txt3':      '#94a3b8',   # Texto terciário
    'border':    '#1e2d40',   # Bordas
    # Terminal
    'term':      '#0d1117',
    'term_green':'#7ee787',
    'term_blue': '#79c0ff',
    'term_ora':  '#ffa657',
    'term_red':  '#ff7b72',
    'term_dim':  '#3d444d',
    'term_txt':  '#c9d1d9',
}

FORMATOS = [
    ('CSV',   '📊', C['blue'],   'Valores separados por vírgula (.csv)'),
    ('Excel', '📗', C['green'],  'Planilha Excel com formatação (.xlsx)'),
    ('XML',   '🗂️', C['purple'], 'Markup estruturado com metadados (.xml)'),
]

EXT_MAP = {'CSV': '.csv', 'Excel': '.xlsx', 'XML': '.xml'}


# ════════════════════════════════════════════════════════════════════════════
class ConversorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Conversor NetCDF  v3.0")
        self.root.configure(bg=C['bg'])

        # Tela cheia no Windows
        try:
            self.root.state('zoomed')
        except Exception:
            self.root.geometry('1300x800')
        self.root.minsize(1000, 660)

        # DPI Awareness (Windows)
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        # ── Estado ────────────────────────────────────────────────────────
        self.arquivos: list[str] = []
        self.processando = False
        self.cancelar = False
        self._tempo_inicio: datetime | None = None
        self._timer_id = None
        self._total_linhas_global = 0

        # ── Variáveis Tk ─────────────────────────────────────────────────
        self.var_arquivos  = tk.StringVar(value='Nenhum arquivo selecionado')
        self.var_destino   = tk.StringVar()
        self.var_status    = tk.StringVar(value='Pronto. Selecione os arquivos para começar.')
        self.var_prog      = tk.DoubleVar(value=0)
        self.var_prog_tot  = tk.DoubleVar(value=0)
        self.var_formato   = tk.StringVar(value='CSV')
        self.var_unificar  = tk.BooleanVar(value=False)
        self.var_resumo    = tk.BooleanVar(value=True)

        self._setup_style()
        self._build_ui()

        if not LIBS_OK:
            self.root.after(600, self._erro_libs)

    # ── ESTILO ──────────────────────────────────────────────────────────────
    def _setup_style(self):
        s = ttk.Style()
        s.theme_use('clam')
        s.configure('File.Horizontal.TProgressbar',
                    troughcolor=C['card2'], background=C['accent'],
                    darkcolor=C['accent'], lightcolor=C['accent_h'],
                    bordercolor=C['border'], thickness=24)
        s.configure('Total.Horizontal.TProgressbar',
                    troughcolor=C['card2'], background=C['blue'],
                    darkcolor=C['blue'],   lightcolor='#60a5fa',
                    bordercolor=C['border'], thickness=10)

    # ════════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # TOP BAR ─────────────────────────────────────────────────────────
        topbar = tk.Frame(self.root, bg=C['accent'], height=60)
        topbar.pack(fill='x')
        topbar.pack_propagate(False)

        tk.Label(topbar, text='🛡️', font=('Segoe UI Emoji', 24),
                 bg=C['accent'], fg='white').pack(side='left', padx=(20, 8))
        tk.Label(topbar, text='CONVERSOR NETCDF',
                 font=('Segoe UI', 17, 'bold'), bg=C['accent'], fg='white').pack(side='left')
        tk.Label(topbar, text='By: Offnen Soluções e Desenvolvimento',
                 font=('Segoe UI', 10), bg=C['accent'], fg='#ffe4c4').pack(side='left', padx=14)

        # Badge versão
        badge = tk.Frame(topbar, bg='#c06010', padx=12, pady=4)
        badge.place(relx=1.0, rely=0.5, anchor='e', x=-18)
        tk.Label(badge, text='v3.0  Desktop', font=('Segoe UI', 9, 'bold'),
                 bg='#c06010', fg='white').pack()

        # CORPO ───────────────────────────────────────────────────────────
        body = tk.Frame(self.root, bg=C['bg'])
        body.pack(fill='both', expand=True)

        # Painel esquerdo (controles)
        self.left = tk.Frame(body, bg=C['bg2'], width=460)
        self.left.pack(side='left', fill='y')
        self.left.pack_propagate(False)

        # Separador vertical
        tk.Frame(body, bg=C['border'], width=2).pack(side='left', fill='y')

        # Painel direito (terminal + estatísticas)
        self.right = tk.Frame(body, bg=C['bg'])
        self.right.pack(side='right', fill='both', expand=True)

        self._build_left()
        self._build_right()

    # ── PAINEL ESQUERDO ──────────────────────────────────────────────────────
    def _build_left(self):
        p = self.left

        # Canvas com scroll para o conteúdo
        canvas = tk.Canvas(p, bg=C['bg2'], highlightthickness=0, bd=0)
        vbar = ttk.Scrollbar(p, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        inner = tk.Frame(canvas, bg=C['bg2'])
        win_id = canvas.create_window((0, 0), window=inner, anchor='nw')

        inner.bind('<Configure>', lambda e: (
            canvas.configure(scrollregion=canvas.bbox('all')),
            canvas.itemconfig(win_id, width=canvas.winfo_width())))
        canvas.bind('<Configure>',
                    lambda e: canvas.itemconfig(win_id, width=e.width))
        # Scroll com roda do mouse
        canvas.bind_all('<MouseWheel>',
                        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))

        PX = 22  # padding horizontal

        # ── SEÇÃO: ARQUIVOS ────────────────────────────────────────────────
        self._secao(inner, '📂  Arquivos de Entrada', px=PX)
        fr = tk.Frame(inner, bg=C['bg2'])
        fr.pack(fill='x', padx=PX, pady=(6, 16))

        entry_arq = tk.Entry(fr, textvariable=self.var_arquivos,
                             font=('Segoe UI', 9), bg=C['card'], fg=C['txt2'],
                             insertbackground=C['txt'], relief='flat', state='readonly',
                             readonlybackground=C['card'])
        entry_arq.pack(side='left', fill='x', expand=True, ipady=10, ipadx=10)
        self._botao(fr, 'Selecionar', self._sel_arquivos, C['accent']
                    ).pack(side='right', padx=(8, 0))

        # ── SEÇÃO: DESTINO ─────────────────────────────────────────────────
        self._secao(inner, '💾  Pasta de Destino', px=PX)
        fr2 = tk.Frame(inner, bg=C['bg2'])
        fr2.pack(fill='x', padx=PX, pady=(6, 16))

        tk.Entry(fr2, textvariable=self.var_destino,
                 font=('Segoe UI', 9), bg=C['card'], fg=C['txt'],
                 insertbackground=C['txt'], relief='flat'
                 ).pack(side='left', fill='x', expand=True, ipady=10, ipadx=10)
        self._botao(fr2, 'Escolher', self._sel_destino, C['accent']
                    ).pack(side='right', padx=(8, 0))

        # ── SEÇÃO: FORMATO ─────────────────────────────────────────────────
        self._secao(inner, '📄  Formato de Saída', px=PX)
        self._fmt_btns: dict = {}

        for fmt, icon, cor, desc in FORMATOS:
            card = tk.Frame(inner, bg=C['card'], cursor='hand2')
            card.pack(fill='x', padx=PX, pady=(0, 8))

            # Ícone
            lbl_icon = tk.Label(card, text=icon, font=('Segoe UI Emoji', 18),
                                bg=C['card'], fg=cor)
            lbl_icon.pack(side='left', padx=(14, 8), pady=12)

            # Textos
            txt_frame = tk.Frame(card, bg=C['card'])
            txt_frame.pack(side='left', fill='x', expand=True, pady=8)
            lbl_nome = tk.Label(txt_frame, text=fmt, font=('Segoe UI', 11, 'bold'),
                                bg=C['card'], fg=C['txt'])
            lbl_nome.pack(anchor='w')
            lbl_desc = tk.Label(txt_frame, text=desc, font=('Segoe UI', 8),
                                bg=C['card'], fg=C['txt2'])
            lbl_desc.pack(anchor='w')

            # Indicador de seleção
            ind = tk.Label(card, text='○', font=('Segoe UI', 16),
                           bg=C['card'], fg=C['txt2'])
            ind.pack(side='right', padx=14)

            self._fmt_btns[fmt] = (card, ind, cor,
                                   [lbl_icon, txt_frame, lbl_nome, lbl_desc])

            for widget in [card, lbl_icon, txt_frame, lbl_nome, lbl_desc, ind]:
                widget.bind('<Button-1>',
                            lambda e, f=fmt, c=cor: self._sel_formato(f, c))

        self._sel_formato('CSV', C['blue'])

        # ── SEÇÃO: OPÇÕES ──────────────────────────────────────────────────
        self._secao(inner, '⚙️  Opções de Processamento', px=PX)
        opts = tk.Frame(inner, bg=C['bg2'])
        opts.pack(fill='x', padx=PX, pady=(6, 16))
        self._check(opts, 'Unificar todos os arquivos em um único arquivo de saída',
                    self.var_unificar)
        self._check(opts, 'Gerar resumo com cálculo de média anual',
                    self.var_resumo)

        # ── SEÇÃO: PROGRESSO ───────────────────────────────────────────────
        self._secao(inner, '📊  Progresso da Conversão', px=PX)

        prog_card = tk.Frame(inner, bg=C['card'], padx=18, pady=16)
        prog_card.pack(fill='x', padx=PX, pady=(6, 16))

        # Barra do arquivo
        row1 = tk.Frame(prog_card, bg=C['card'])
        row1.pack(fill='x', pady=(0, 5))
        tk.Label(row1, text='Arquivo atual', font=('Segoe UI', 9),
                 fg=C['txt2'], bg=C['card']).pack(side='left')
        self.lbl_pct = tk.Label(row1, text='0%', font=('Segoe UI', 14, 'bold'),
                                fg=C['accent'], bg=C['card'])
        self.lbl_pct.pack(side='right')

        self.pb_arq = ttk.Progressbar(prog_card, variable=self.var_prog,
                                      maximum=100, mode='determinate',
                                      style='File.Horizontal.TProgressbar')
        self.pb_arq.pack(fill='x', pady=(0, 14))

        # Barra total
        row2 = tk.Frame(prog_card, bg=C['card'])
        row2.pack(fill='x', pady=(0, 5))
        tk.Label(row2, text='Total do lote', font=('Segoe UI', 9),
                 fg=C['txt2'], bg=C['card']).pack(side='left')
        self.lbl_tot_pct = tk.Label(row2, text='0%', font=('Segoe UI', 9, 'bold'),
                                    fg=C['blue'], bg=C['card'])
        self.lbl_tot_pct.pack(side='right')

        self.pb_tot = ttk.Progressbar(prog_card, variable=self.var_prog_tot,
                                      maximum=100, mode='determinate',
                                      style='Total.Horizontal.TProgressbar')
        self.pb_tot.pack(fill='x', pady=(0, 12))

        self.lbl_status = tk.Label(prog_card, textvariable=self.var_status,
                                   font=('Segoe UI', 9), fg=C['txt2'],
                                   bg=C['card'], wraplength=380, justify='left')
        self.lbl_status.pack(fill='x')

        # ── BOTÕES ─────────────────────────────────────────────────────────
        btn_area = tk.Frame(inner, bg=C['bg2'])
        btn_area.pack(fill='x', padx=PX, pady=(10, 24))

        self.btn_conv = tk.Button(
            btn_area, text='▶   INICIAR CONVERSÃO',
            command=self._iniciar,
            font=('Segoe UI', 13, 'bold'),
            bg=C['green'], fg='white',
            activebackground='#16a34a', activeforeground='white',
            cursor='hand2', relief='flat', pady=16, borderwidth=0)
        self.btn_conv.pack(fill='x', pady=(0, 8))

        self.btn_cancel = tk.Button(
            btn_area, text='⏹   CANCELAR',
            command=self._cancelar,
            font=('Segoe UI', 11),
            bg=C['card'], fg=C['txt2'],
            activebackground=C['red'], activeforeground='white',
            cursor='hand2', relief='flat', pady=11, borderwidth=0,
            state='disabled')
        self.btn_cancel.pack(fill='x')

        # Rodapé
        footer = tk.Frame(self.left, bg=C['border'], height=30)
        footer.pack(fill='x', side='bottom')
        footer.pack_propagate(False)
        tk.Label(footer, text='By: Offnen Soluções e Desenvolvimento',
                 font=('Segoe UI', 8), fg=C['txt2'], bg=C['border']
                 ).place(relx=0.5, rely=0.5, anchor='center')

    # ── PAINEL DIREITO ───────────────────────────────────────────────────────
    def _build_right(self):
        r = self.right

        # Barra de estatísticas
        stats_bar = tk.Frame(r, bg=C['card2'], height=62)
        stats_bar.pack(fill='x')
        stats_bar.pack_propagate(False)

        self.st_files = self._stat(stats_bar, '0',     'arquivos')
        self.st_size  = self._stat(stats_bar, '0 MB',  'tamanho total')
        self.st_fmt   = self._stat(stats_bar, 'CSV',   'formato')
        self.st_time  = self._stat(stats_bar, '00:00', 'tempo decorrido')
        self.st_lines = self._stat(stats_bar, '0',     'linhas exportadas')

        # Cabeçalho do terminal
        term_head = tk.Frame(r, bg='#161b22', height=40)
        term_head.pack(fill='x')
        term_head.pack_propagate(False)

        # Dots estilo macOS
        dots = tk.Frame(term_head, bg='#161b22')
        dots.place(relx=0, rely=0.5, anchor='w', x=16)
        for dot_col in ['#ff5f57', '#febc2e', '#28c840']:
            tk.Frame(dots, bg=dot_col, width=14, height=14).pack(side='left', padx=3)

        tk.Label(term_head,
                 text='⚡  LOG DE CONVERSÃO  —  TEMPO REAL',
                 font=('Consolas', 9, 'bold'), fg='#8b949e',
                 bg='#161b22').place(relx=0.5, rely=0.5, anchor='center')

        btn_clear = tk.Button(term_head, text='Limpar log',
                              font=('Segoe UI', 8),
                              bg='#21262d', fg='#8b949e',
                              activebackground='#30363d', activeforeground='white',
                              relief='flat', borderwidth=0, padx=10, pady=3,
                              cursor='hand2', command=self._limpar_log)
        btn_clear.place(relx=1.0, rely=0.5, anchor='e', x=-16)

        # Área do terminal
        self.term = scrolledtext.ScrolledText(
            r,
            font=('Consolas', 10),
            bg=C['term'], fg=C['term_txt'],
            insertbackground=C['term_txt'],
            relief='flat', borderwidth=0,
            wrap='word', state='disabled', cursor='arrow',
            selectbackground='#264f78', selectforeground='white',
            padx=12, pady=8)
        self.term.pack(fill='both', expand=True)

        # Tags de cor
        self.term.tag_configure('success', foreground=C['term_green'])
        self.term.tag_configure('error',   foreground=C['term_red'])
        self.term.tag_configure('warning', foreground=C['term_ora'])
        self.term.tag_configure('info',    foreground=C['term_blue'])
        self.term.tag_configure('dim',     foreground=C['term_dim'])
        self.term.tag_configure('accent',  foreground=C['accent'],
                                font=('Consolas', 10, 'bold'))
        self.term.tag_configure('header',  foreground='#e6edf3',
                                font=('Consolas', 11, 'bold'))
        self.term.tag_configure('ts',      foreground='#3d444d',
                                font=('Consolas', 9))
        self.term.tag_configure('xml',     foreground='#a5f3a5')
        self.term.tag_configure('excel',   foreground='#6ee7b7')
        self.term.tag_configure('csv',     foreground=C['term_blue'])

        self._log_boas_vindas()

    # ── HELPERS DE UI ────────────────────────────────────────────────────────
    def _secao(self, parent, texto, px=0):
        f = tk.Frame(parent, bg=C['bg2'])
        f.pack(fill='x', padx=px, pady=(16, 0))
        tk.Label(f, text=texto, font=('Segoe UI', 10, 'bold'),
                 fg=C['txt'], bg=C['bg2']).pack(anchor='w')
        tk.Frame(f, bg=C['accent'], height=2).pack(fill='x', pady=(3, 0))

    def _botao(self, parent, texto, cmd, cor):
        return tk.Button(parent, text=texto, command=cmd,
                         font=('Segoe UI', 9, 'bold'), bg=cor, fg='white',
                         activebackground=C['accent_h'], activeforeground='white',
                         cursor='hand2', relief='flat', padx=14, pady=8, borderwidth=0)

    def _check(self, parent, texto, var):
        tk.Checkbutton(
            parent, text=texto, variable=var,
            font=('Segoe UI', 9), bg=C['bg2'], fg=C['txt'],
            activebackground=C['bg2'], activeforeground=C['accent'],
            selectcolor=C['card'], cursor='hand2'
        ).pack(anchor='w', pady=3)

    def _stat(self, parent, valor, label):
        f = tk.Frame(parent, bg=C['card2'])
        f.pack(side='left', padx=20, fill='y', pady=10)
        v = tk.Label(f, text=valor, font=('Segoe UI', 14, 'bold'),
                     fg=C['accent'], bg=C['card2'])
        v.pack()
        tk.Label(f, text=label, font=('Segoe UI', 7),
                 fg=C['txt2'], bg=C['card2']).pack()
        return v

    def _sel_formato(self, fmt: str, cor: str):
        self.var_formato.set(fmt)
        self.st_fmt.configure(text=fmt)
        for f, (card, ind, c, widgets) in self._fmt_btns.items():
            if f == fmt:
                card.configure(bg=c)
                ind.configure(text='●', fg='white', bg=c)
                for w in widgets:
                    try:
                        w.configure(bg=c)
                        if isinstance(w, tk.Label):
                            w.configure(fg='white')
                    except Exception:
                        pass
            else:
                card.configure(bg=C['card'])
                ind.configure(text='○', fg=C['txt2'], bg=C['card'])
                for w in widgets:
                    try:
                        w.configure(bg=C['card'])
                        if isinstance(w, tk.Label):
                            w.configure(fg=C['txt'] if 'bold' in str(
                                w.cget('font')) else C['txt2'])
                    except Exception:
                        pass

    # ── LOG / TERMINAL ───────────────────────────────────────────────────────
    def _log(self, msg: str, tag: str = None):
        def _do():
            self.term.configure(state='normal')
            ts = datetime.now().strftime('%H:%M:%S.%f')[:-3]
            if not msg:
                self.term.insert('end', '\n')
            else:
                self.term.insert('end', f'[{ts}] ', 'ts')
                self.term.insert('end', f'{msg}\n', tag or 'success')
            self.term.see('end')
            self.term.configure(state='disabled')
        self.root.after(0, _do)

    def _limpar_log(self):
        self.term.configure(state='normal')
        self.term.delete('1.0', 'end')
        self.term.configure(state='disabled')
        self._log_boas_vindas()

    def _log_boas_vindas(self):
        self._log('╔══════════════════════════════════════════════════════════════╗', 'dim')
        self._log('║   CONVERSOR NETCDF  v3.0                                     ║', 'header')
        self._log('║   Saída:  📊 CSV  │  📗 Excel (XLSX)  │  🗂️ XML              ║', 'dim')
        self._log('║   Capacidade: arquivos de 3 GB, 4 GB ou mais                 ║', 'dim')
        self._log('╚══════════════════════════════════════════════════════════════╝', 'dim')
        self._log('')
        self._log('  ①  Selecione os arquivos NetCDF (.nc)', 'info')
        self._log('  ②  Escolha o formato de saída', 'info')
        self._log('  ③  Clique em INICIAR CONVERSÃO', 'info')
        self._log('')
        self._log('  O log de progresso em tempo real aparecerá aqui.', 'dim')
        self._log('')

    # ── SELEÇÃO DE ARQUIVOS ─────────────────────────────────────────────────
    def _sel_arquivos(self):
        res = filedialog.askopenfilenames(
            title='Selecione arquivos NetCDF (.nc)',
            filetypes=[('NetCDF', '*.nc'), ('Todos os arquivos', '*.*')],
            parent=self.root)
        if not res:
            return

        self.arquivos = list(res)
        n = len(self.arquivos)

        if n == 1:
            self.var_arquivos.set(Path(self.arquivos[0]).name)
            self.var_destino.set(str(Path(self.arquivos[0]).parent))
        else:
            self.var_arquivos.set(f'{n} arquivos selecionados')
            self.var_destino.set(str(Path(self.arquivos[0]).parent))

        total_bytes = sum(os.path.getsize(f) for f in self.arquivos)
        sz = self._fmt_size(total_bytes)

        self.st_files.configure(text=str(n))
        self.st_size.configure(text=sz)
        self.var_status.set(f'✅ {n} arquivo(s) prontos — {sz}')

        self._log(f'Selecionados: {n} arquivo(s)  —  {sz}', 'info')
        for f in self.arquivos:
            self._log(f'  📄 {Path(f).name}  ({self._fmt_size(os.path.getsize(f))})', 'dim')
        self._log('')

    def _sel_destino(self):
        d = filedialog.askdirectory(title='Escolha a pasta de destino', parent=self.root)
        if d:
            self.var_destino.set(d)
            self._log(f'Pasta de destino: {d}', 'info')

    def _fmt_size(self, b: int) -> str:
        if b < 1024:       return f'{b} B'
        if b < 1024**2:    return f'{b/1024:.1f} KB'
        if b < 1024**3:    return f'{b/1024**2:.0f} MB'
        return f'{b/1024**3:.2f} GB'

    # ── INICIAR CONVERSÃO ────────────────────────────────────────────────────
    def _iniciar(self):
        if not LIBS_OK:
            self._erro_libs()
            return
        if not self.arquivos:
            messagebox.showwarning('Aviso', 'Selecione pelo menos um arquivo NetCDF.')
            return
        destino = self.var_destino.get()
        if not destino or not os.path.exists(destino):
            messagebox.showwarning('Aviso', 'Escolha uma pasta de destino válida.')
            return
        fmt = self.var_formato.get()
        if fmt == 'Excel' and not EXCEL_OK:
            messagebox.showerror(
                'Dependência ausente',
                'openpyxl não está instalado.\nExecute: pip install openpyxl')
            return

        # Atualizar UI
        self.processando = True
        self.cancelar = False
        self._total_linhas_global = 0
        self.btn_conv.configure(state='disabled', bg=C['card'], fg=C['txt2'])
        self.btn_cancel.configure(state='normal', bg=C['red'], fg='white')
        self.var_prog.set(0)
        self.var_prog_tot.set(0)
        self.lbl_pct.configure(text='0%')
        self.lbl_tot_pct.configure(text='0%')
        self.st_lines.configure(text='0')

        self._tempo_inicio = datetime.now()
        self._iniciar_timer()

        self._log('')
        self._log('━' * 62, 'dim')
        self._log(f'  INICIANDO CONVERSÃO  ▶  {datetime.now():%d/%m/%Y %H:%M:%S}', 'accent')
        self._log(f'  Formato: {fmt}   |   Arquivos: {len(self.arquivos)}', 'info')
        self._log(f'  Destino: {destino}', 'dim')
        self._log('━' * 62, 'dim')
        self._log('')

        t = threading.Thread(
            target=self._processar_lote,
            args=(self.arquivos, destino, fmt),
            daemon=True)
        t.start()

    def _iniciar_timer(self):
        def tick():
            if self._tempo_inicio and self.processando:
                elapsed = int((datetime.now() - self._tempo_inicio).total_seconds())
                m, s = divmod(elapsed, 60)
                self.st_time.configure(text=f'{m:02d}:{s:02d}')
                self._timer_id = self.root.after(1000, tick)
        self._timer_id = self.root.after(1000, tick)

    def _cancelar(self):
        if self.processando:
            self.cancelar = True
            self.var_status.set('⏳ Cancelando...')
            self._log('Cancelamento solicitado pelo usuário.', 'warning')

    def _atualizar_progresso(self, arq: float, tot: float, msg: str):
        self.var_prog.set(arq)
        self.var_prog_tot.set(tot)
        self.lbl_pct.configure(text=f'{int(arq)}%')
        self.lbl_tot_pct.configure(text=f'{int(tot)}%')
        self.var_status.set(msg)
        self.root.update_idletasks()

    # ── PROCESSAMENTO EM LOTE ────────────────────────────────────────────────
    def _processar_lote(self, arquivos: list, destino: str, fmt: str):
        total = len(arquivos)
        sucessos = 0
        erros: list[str] = []
        linhas_total = 0
        unificar = self.var_unificar.get()
        gerar_resumo = self.var_resumo.get()
        ext = EXT_MAP.get(fmt, '.csv')

        arquivo_unico = None
        if unificar:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            arquivo_unico = str(Path(destino) / f'NetCDF_Unificado_{ts}{ext}')

        stats_global = None

        for idx, entrada in enumerate(arquivos):
            if self.cancelar:
                break

            nome = Path(entrada).name
            prog_base = (idx / total) * 100

            self._log(f'[{idx+1}/{total}] ► {nome}', 'accent')

            saida = arquivo_unico if unificar else str(
                Path(destino) / f'{Path(entrada).stem}{ext}')
            append = unificar and idx > 0

            try:
                res = self._converter_arquivo(
                    entrada=entrada,
                    saida=saida,
                    fmt=fmt,
                    prog_offset=prog_base,
                    prog_peso=100 / total,
                    append=append,
                    write_header=(not unificar or idx == 0),
                    calcular_stats=gerar_resumo)

                if res['linhas'] > 0:
                    sucessos += 1
                    linhas_total += res['linhas']
                    self.root.after(0, lambda n=linhas_total:
                        self.st_lines.configure(text=f'{n:,}'))
                    self._log(
                        f'  ✓  {res["linhas"]:,} linhas  →  {Path(saida).name}',
                        'success')

                    if gerar_resumo and res['stats'] is not None:
                        if stats_global is None:
                            stats_global = res['stats']
                        else:
                            stats_global = stats_global.add(res['stats'], fill_value=0)

                        if not unificar:
                            rp = str(Path(destino) / f'{Path(entrada).stem}_resumo_anual{ext}')
                            self._salvar_resumo(res['stats'], rp, fmt)
                            self._log(f'  📈 Resumo anual → {Path(rp).name}', 'info')

            except Exception as e:
                erros.append(f'{nome}: {e}')
                self._log(f'  ✗  Erro: {e}', 'error')

            self._log('')

        # Resumo global
        if gerar_resumo and unificar and stats_global is not None:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            rp = str(Path(destino) / f'NetCDF_Resumo_Global_{ts}{ext}')
            self._salvar_resumo(stats_global, rp, fmt)
            self._log(f'📈 Resumo global → {Path(rp).name}', 'info')

        # Finalizar
        if self.cancelar:
            self._finalizar(False, f'❌ Cancelado — {sucessos}/{total} convertidos')
        elif erros and sucessos == 0:
            self._finalizar(False,
                            f'❌ Falha em todos os {total} arquivo(s).\nPrimeiro erro: {erros[0]}')
        else:
            msg = f'✅ {sucessos}/{total} arquivo(s) convertidos para {fmt}!'
            if linhas_total:
                msg += f'  ({linhas_total:,} linhas no total)'
            self._finalizar(True, msg)

    # ── CONVERTER ARQUIVO INDIVIDUAL ─────────────────────────────────────────
    def _converter_arquivo(self, entrada, saida, fmt,
                           prog_offset, prog_peso,
                           append, write_header, calcular_stats):
        self._log('  → Abrindo dataset NetCDF...', 'dim')
        self._atualizar_progresso(0, prog_offset, f'Abrindo {Path(entrada).name}...')

        ds = xr.open_dataset(entrada)

        if self.cancelar:
            ds.close()
            return {'linhas': 0, 'stats': None}

        dims    = list(ds.dims.keys())
        dim0    = dims[0]
        tam     = ds.dims[dim0]
        vars_l  = list(ds.data_vars.keys())
        total_pts = 1
        for d in ds.dims.values():
            total_pts *= d

        self._log(f'  → Dimensões: {dict(ds.dims)}', 'dim')
        self._log(f'  → Variáveis: {vars_l}', 'dim')

        chunk = max(1, tam // (200 if total_pts > 50_000_000 else 20))
        modo  = 'grande' if total_pts > 50_000_000 else 'normal'
        self._log(f'  → Modo: {modo}  |  Chunk: {chunk}  |  Pts: {total_pts:,}', 'dim')

        var_precip = 'pr' if 'pr' in ds.data_vars else vars_l[0]

        total_linhas = 0
        df_stats     = None
        primeiro     = write_header
        chunks_data: list[pd.DataFrame] = []

        n_chunks = (tam + chunk - 1) // chunk

        for ci, i in enumerate(range(0, tam, chunk)):
            if self.cancelar:
                ds.close()
                return {'linhas': 0, 'stats': None}

            fim = min(i + chunk, tam)
            pct_arq = (i / tam) * 100
            pct_tot = prog_offset + prog_peso * (i / tam)

            self._atualizar_progresso(
                pct_arq, pct_tot,
                f'{Path(entrada).name}: {int(pct_arq)}%')

            # Log a cada 20%
            if ci % max(1, n_chunks // 5) == 0:
                tag = {'CSV': 'csv', 'Excel': 'excel', 'XML': 'xml'}.get(fmt, 'dim')
                self._log(
                    f'  → Chunk {ci+1}/{n_chunks}  —  {int(pct_arq)}%  '
                    f'({total_linhas:,} linhas até agora)', tag)

            subset = ds.isel({dim0: slice(i, fim)})
            df = subset.to_dataframe().reset_index()
            df = df.replace([np.inf, -np.inf], np.nan)

            # Estatísticas anuais
            if calcular_stats:
                try:
                    col_t = 'time' if 'time' in df.columns else df.columns[0]
                    df[col_t] = pd.to_datetime(df[col_t])
                    col_lat = next((c for c in ['latitude', 'lat'] if c in df.columns), None)
                    col_lon = next((c for c in ['longitude', 'lon'] if c in df.columns), None)
                    grp = [df[col_t].dt.year]
                    if col_lat: grp.append(col_lat)
                    if col_lon: grp.append(col_lon)
                    cs = df.groupby(grp)[var_precip].agg(['sum', 'count'])
                    df_stats = cs if df_stats is None else df_stats.add(cs, fill_value=0)
                except Exception as e:
                    self._log(f'  ⚠ Estatísticas: {e}', 'warning')

            # Escrita
            if fmt == 'CSV':
                mode = 'a' if (append or not primeiro) else 'w'
                df.to_csv(saida, index=False, encoding='utf-8-sig',
                          mode=mode, header=primeiro)
                primeiro = False
            else:
                chunks_data.append(df.copy())

            total_linhas += len(df)
            del df, subset
            gc.collect()

        ds.close()
        gc.collect()

        # ── Excel ──────────────────────────────────────────────────────────
        if fmt == 'Excel' and chunks_data:
            self._log('  → Montando planilha Excel...', 'info')
            df_all = pd.concat(chunks_data, ignore_index=True)
            del chunks_data

            if append and os.path.exists(saida):
                existing = pd.read_excel(saida, engine='openpyxl')
                df_all = pd.concat([existing, df_all], ignore_index=True)

            df_all.to_excel(saida, index=False, engine='openpyxl')

            # Formatação visual
            try:
                wb = load_workbook(saida)
                ws = wb.active
                h_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F',
                                     fill_type='solid')
                h_font = Font(color='FFFFFF', bold=True, size=10)
                h_alig = Alignment(horizontal='center', vertical='center')
                thin   = Side(style='thin', color='2d5a87')
                border = Border(bottom=thin)
                for cell in ws[1]:
                    cell.fill   = h_fill
                    cell.font   = h_font
                    cell.alignment = h_alig
                    cell.border = border
                ws.freeze_panes = 'A2'
                # Auto-width (limitado a 50 chars)
                for col in ws.columns:
                    max_w = max(len(str(c.value or '')) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 50)
                wb.save(saida)
            except Exception as e:
                self._log(f'  ⚠ Formatação Excel: {e}', 'warning')

            self._log(f'  → Excel gravado: {len(df_all):,} linhas', 'excel')
            del df_all
            gc.collect()

        # ── XML ─────────────────────────────────────────────────────────────
        elif fmt == 'XML' and chunks_data:
            self._log('  → Gerando arquivo XML...', 'info')
            df_all = pd.concat(chunks_data, ignore_index=True)
            del chunks_data
            self._escrever_xml(df_all, saida, append)
            self._log(f'  → XML gravado: {len(df_all):,} registros', 'xml')
            del df_all
            gc.collect()

        return {'linhas': total_linhas, 'stats': df_stats}

    # ── ESCREVER XML ─────────────────────────────────────────────────────────
    def _escrever_xml(self, df: 'pd.DataFrame', path: str, append: bool = False):
        root_el = ET.Element('NetCDF_Dataset')
        root_el.set('gerado_em',        datetime.now().isoformat())
        root_el.set('total_registros',  str(len(df)))
        root_el.set('fonte',            'Conversor NetCDF v3.0')
        root_el.set('versao',           '3.0')

        meta = ET.SubElement(root_el, 'metadados')
        cols_el = ET.SubElement(meta, 'colunas')
        for col in df.columns:
            c_el = ET.SubElement(cols_el, 'coluna')
            c_el.set('nome', str(col))
            c_el.set('tipo', str(df[col].dtype))

        dados = ET.SubElement(root_el, 'dados')
        for _, row in df.iterrows():
            reg = ET.SubElement(dados, 'registro')
            for col in df.columns:
                val = row[col]
                tag_name = (str(col)
                            .replace(' ', '_')
                            .replace('/', '_')
                            .replace('.', '_'))
                if not tag_name[0].isalpha():
                    tag_name = 'f_' + tag_name
                el = ET.SubElement(reg, tag_name)
                el.text = '' if pd.isna(val) else str(val)

        tree = ET.ElementTree(root_el)
        try:
            ET.indent(tree, space='  ')
        except AttributeError:
            pass  # Python < 3.9 — sem indentação automática
        tree.write(path, encoding='utf-8', xml_declaration=True)

    # ── SALVAR RESUMO ────────────────────────────────────────────────────────
    def _salvar_resumo(self, df_stats: 'pd.DataFrame', path: str, fmt: str):
        try:
            resumo = df_stats.copy()
            resumo['media_anual'] = resumo['sum'] / resumo['count']
            resumo = resumo.reset_index()
            if 'time' in resumo.columns:
                resumo = resumo.rename(columns={'time': 'ano'})
            if 'sum' in resumo.columns:
                resumo = resumo.rename(columns={'sum': 'precip_acumulada_anual'})

            if fmt == 'CSV':
                resumo.to_csv(path, index=False, encoding='utf-8-sig')
            elif fmt == 'Excel':
                resumo.to_excel(path, index=False, engine='openpyxl')
            elif fmt == 'XML':
                self._escrever_xml(resumo, path)
        except Exception as e:
            self._log(f'  ⚠ Erro ao salvar resumo: {e}', 'warning')

    # ── FINALIZAR ────────────────────────────────────────────────────────────
    def _finalizar(self, sucesso: bool, msg: str):
        self.processando = False
        if self._timer_id:
            self.root.after_cancel(self._timer_id)

        self.var_prog.set(100 if sucesso else 0)
        self.var_prog_tot.set(100 if sucesso else 0)
        self.lbl_pct.configure(text='100%' if sucesso else '0%')
        self.lbl_tot_pct.configure(text='100%' if sucesso else '0%')
        self.var_status.set(msg)

        self.btn_conv.configure(state='normal', bg=C['green'], fg='white')
        self.btn_cancel.configure(state='disabled', bg=C['card'], fg=C['txt2'])

        self._log('')
        self._log('━' * 62, 'dim')
        self._log(f'  {msg}', 'success' if sucesso else 'error')
        if self._tempo_inicio:
            elapsed = int((datetime.now() - self._tempo_inicio).total_seconds())
            m, s = divmod(elapsed, 60)
            self._log(f'  Tempo total: {m:02d}m {s:02d}s', 'info')
        self._log('━' * 62, 'dim')

        self.lbl_status.configure(fg=C['green'] if sucesso else C['red'])

        if sucesso:
            self.root.after(200, lambda: messagebox.showinfo(
                'Conversão Concluída! 🎉',
                f'{msg}\n\nArquivos salvos em:\n{self.var_destino.get()}'))
        elif 'cancelad' not in msg.lower():
            self.root.after(200, lambda: messagebox.showerror('Erro', msg))

    def _erro_libs(self):
        messagebox.showerror(
            'Bibliotecas não encontradas',
            f'Erro ao carregar biblioteca:\n{MISSING_LIB}\n\n'
            'Execute no terminal:\n'
            'pip install xarray netcdf4 pandas numpy openpyxl')


# ── ENTRY POINT ─────────────────────────────────────────────────────────────
def main():
    root = tk.Tk()
    try:
        root.iconbitmap('icon.ico')
    except Exception:
        pass
    ConversorApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()

"""
Conversor NetCDF Local - Defesa Civil Araruna
Script para processar arquivos NetCDF grandes diretamente no computador.

Uso:
    python converter_local.py arquivo.nc [--formato xlsx|csv] [--output pasta_saida]

Exemplos:
    python converter_local.py dados.nc
    python converter_local.py dados.nc --formato csv
    python converter_local.py dados.nc --output C:\Meus_Dados
"""

import argparse
import sys
import os
from pathlib import Path
from datetime import datetime

import xarray as xr
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Caminho para a logo da prefeitura
LOGO_PATH = Path(__file__).parent.parent / "public" / "images" / "logo-prefeitura.png"


def processar_netcdf(caminho_arquivo: str) -> tuple[xr.Dataset, dict]:
    """Abrir e processar arquivo NetCDF"""
    print(f"📂 Abrindo arquivo: {caminho_arquivo}")
    print("   Isso pode demorar para arquivos grandes...")
    
    ds = xr.open_dataset(caminho_arquivo)
    
    metadados = {
        "variaveis": list(ds.data_vars.keys()),
        "dimensoes": {dim: int(size) for dim, size in ds.dims.items()},
        "coordenadas": list(ds.coords.keys()),
    }
    
    print(f"✅ Arquivo carregado!")
    print(f"   Variáveis: {', '.join(metadados['variaveis'])}")
    print(f"   Dimensões: {metadados['dimensoes']}")
    
    return ds, metadados


def dataset_para_dataframe(ds: xr.Dataset) -> pd.DataFrame:
    """Converter Dataset xarray para DataFrame pandas"""
    print("🔄 Convertendo para tabela...")
    
    df = ds.to_dataframe().reset_index()
    df = df.replace([np.inf, -np.inf], np.nan)
    
    print(f"   Linhas: {len(df):,}")
    print(f"   Colunas: {len(df.columns)}")
    
    return df


def criar_excel_com_logo(df: pd.DataFrame, nome_arquivo: str, caminho_saida: str) -> str:
    """Criar arquivo Excel com logo da prefeitura"""
    print("📊 Criando arquivo Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    start_row = 1
    
    # Adicionar logo se existir
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width = 80
            img.height = 90
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 70
            start_row = 6
            print("   ✅ Logo adicionada!")
        except Exception as e:
            print(f"   ⚠️ Não foi possível adicionar logo: {e}")
    
    # Título do documento
    ws.cell(row=start_row, column=1, value="PREFEITURA MUNICIPAL DE ARARUNA/PB")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=min(len(df.columns), 6))
    
    ws.cell(row=start_row + 1, column=1, value="DEFESA CIVIL - Dados Meteorológicos")
    ws.cell(row=start_row + 1, column=1).font = Font(bold=True, size=12, color="E87722")
    
    ws.cell(row=start_row + 2, column=1, value=f"Arquivo: {nome_arquivo}")
    ws.cell(row=start_row + 2, column=1).font = Font(italic=True, size=10)
    
    ws.cell(row=start_row + 3, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=start_row + 3, column=1).font = Font(italic=True, size=10)
    
    data_start_row = start_row + 5
    
    # Adicionar cabeçalhos
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=str(column))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adicionar dados (com progresso)
    total_rows = len(df)
    print(f"   Escrevendo {total_rows:,} linhas...")
    
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row + 1):
        if row_idx % 10000 == 0:
            progress = ((row_idx - data_start_row) / total_rows) * 100
            print(f"   Progresso: {progress:.1f}%", end='\r')
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if pd.isna(value):
                cell.value = ""
            elif isinstance(value, (np.integer, np.floating)):
                cell.value = float(value)
            elif isinstance(value, np.datetime64):
                cell.value = pd.Timestamp(value).to_pydatetime()
            else:
                cell.value = str(value)
            
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
    
    print("   Progresso: 100.0%")
    
    # Salvar
    arquivo_saida = os.path.join(caminho_saida, nome_arquivo.replace('.nc', '.xlsx'))
    wb.save(arquivo_saida)
    
    return arquivo_saida


def criar_csv(df: pd.DataFrame, nome_arquivo: str, caminho_saida: str) -> str:
    """Criar arquivo CSV"""
    print("📊 Criando arquivo CSV...")
    
    arquivo_saida = os.path.join(caminho_saida, nome_arquivo.replace('.nc', '.csv'))
    df.to_csv(arquivo_saida, index=False, encoding='utf-8-sig')
    
    return arquivo_saida


def main():
    parser = argparse.ArgumentParser(
        description='Conversor NetCDF - Defesa Civil Araruna',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
    python converter_local.py dados.nc
    python converter_local.py dados.nc --formato csv
    python converter_local.py dados.nc --output C:\\Meus_Dados
        """
    )
    
    parser.add_argument('arquivo', help='Caminho do arquivo NetCDF (.nc)')
    parser.add_argument('--formato', choices=['xlsx', 'csv'], default='xlsx',
                       help='Formato de saída (padrão: xlsx)')
    parser.add_argument('--output', '-o', default=None,
                       help='Pasta de saída (padrão: mesma pasta do arquivo)')
    
    args = parser.parse_args()
    
    # Verificar se arquivo existe
    if not os.path.exists(args.arquivo):
        print(f"❌ Erro: Arquivo não encontrado: {args.arquivo}")
        sys.exit(1)
    
    if not args.arquivo.endswith('.nc'):
        print(f"❌ Erro: Arquivo deve ter extensão .nc")
        sys.exit(1)
    
    # Pasta de saída
    if args.output:
        caminho_saida = args.output
        os.makedirs(caminho_saida, exist_ok=True)
    else:
        caminho_saida = os.path.dirname(os.path.abspath(args.arquivo))
    
    nome_arquivo = os.path.basename(args.arquivo)
    
    print()
    print("=" * 60)
    print("   DEFESA CIVIL ARARUNA - Conversor NetCDF")
    print("=" * 60)
    print()
    
    try:
        # Processar
        ds, metadados = processar_netcdf(args.arquivo)
        df = dataset_para_dataframe(ds)
        ds.close()
        
        # Converter
        if args.formato == 'xlsx':
            arquivo_saida = criar_excel_com_logo(df, nome_arquivo, caminho_saida)
        else:
            arquivo_saida = criar_csv(df, nome_arquivo, caminho_saida)
        
        print()
        print("=" * 60)
        print(f"✅ Conversão concluída!")
        print(f"📁 Arquivo salvo em: {arquivo_saida}")
        print("=" * 60)
        
    except Exception as e:
        print()
        print(f"❌ Erro durante a conversão: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
